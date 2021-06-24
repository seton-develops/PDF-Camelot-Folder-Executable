'''
Created on Jun 9, 2021
@description: Converts PDF to Pandas df and writes data to a specified csv file
@author: Sean Tonthat
'''


import os
import camelot
import openpyxl
from time import time

#path = 'C:\\Users\\Sean\\Documents\\PDF_to_CSV\\testh700'

def convert(in_path, out_path, edge_tol2 = 500, row_tol2 = 2):
    
    start = time()
    combined_wbk = ""
    final_path = ""

    for root, _, files in os.walk(in_path, topdown=False):
        for name in files:
            file_path = str(os.path.join(root, name)) #get full path
            print(file_path)
        
            #get tables
            tables = camelot.read_pdf(file_path, pages = 'all', flavor = 'stream', edge_tol=edge_tol2, row_tol = row_tol2 )
            #tables = camelot.read_pdf(file_path, pages = 'all', flavor = 'lattice',line_scale = 40)
        
            #EXPORT
            new_path = out_path + '\\' + str(name).replace(".pdf","") + "-new.xlsx"
            print("EXPORTING to " + new_path)
            tables.export(new_path, f='excel')
        
            wbk = openpyxl.load_workbook(new_path)
            combined_wbk = openpyxl.Workbook()
            ws = combined_wbk.active
            ws.title = "combined"

            print("COMBINING SHEETS")
            for i in range(len(wbk.sheetnames)):
                for row in wbk[wbk.sheetnames[i]].rows:
                    combined_wbk['combined'].append([j.value for j in row])
                    #add some space to separate table pages
                combined_wbk['combined'].append([])
                combined_wbk['combined'].append([])
                combined_wbk['combined'].append([])
                wbk.close()
            
            #save combined workbook
            final_path = out_path+ "\\" + str(name).replace(".pdf","") + "-combined.xlsx"
            combined_wbk.save(final_path)
            combined_wbk.close()
            print("Combined Successfully")
            print("new file in " , final_path)
        
            #print time it took to complete process
            print("time elapsed: ", time() - start)











'''
start = time()
combined_wbk = ""
final_path = ""

for root, directories, files in os.walk(path, topdown=False):
    for name in files:
        file_path = str(os.path.join(root, name)) #get full path
        print(file_path)
        
        #get tables
        tables = camelot.read_pdf(file_path, pages = 'all', flavor = 'stream', edge_tol=500, row_tol = 2 )
        #tables = camelot.read_pdf(file_path, pages = 'all', flavor = 'lattice',line_scale = 40)
        
        #EXPORT
        new_path = "C:\\Users\\Sean\\Documents\\PDF_to_CSV\\H700_new\\" + str(name).replace(".pdf","") + "-new.xlsx"
        print("EXPORTING to " + new_path)
        tables.export(new_path, f='excel')
        
        wbk = openpyxl.load_workbook(new_path)
        combined_wbk = openpyxl.Workbook()
        ws = combined_wbk.active
        ws.title = "combined"

        print("COMBINING SHEETS")
        for i in range(len(wbk.sheetnames)):
            for row in wbk[wbk.sheetnames[i]].rows:
                combined_wbk['combined'].append([j.value for j in row])
            #add some space to separate table pages
            combined_wbk['combined'].append([])
            combined_wbk['combined'].append([])
            combined_wbk['combined'].append([])
            wbk.close() #POSSIBLE ERROR HERE
            
        #save combined workbook
        final_path = "C:\\Users\\Sean\\Documents\\PDF_to_CSV\\H700_new\\" + str(name).replace(".pdf","") + "-combined.xlsx"
        combined_wbk.save(final_path)
        combined_wbk.close()
        print("Combined Successfully")
        print("new file in " , final_path)
        
        #print time it took to complete process
        print("time elapsed: ", time() - start)
'''        


